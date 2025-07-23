#!/usr/bin/env python3
"""
Excel File Reordering Script

This script reorders rows in an Excel file based on IDs from another Excel file.
It matches IDs while ignoring additional text and preserves formatting.

Usage:
    python excel_reorder.py

Requirements:
    pip install pandas openpyxl

Author: Generated for Jakob's thesis data processing
"""

import pandas as pd
import re
from pathlib import Path
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse


def extract_id(cell_value):
    """
    Extract ID from cell value, ignoring additional text.
    
    Handles special cases like:
    - "Col-0_D - Eu" -> "Col_0_D"
    - "Ler-D_E - Eu" -> "Ler_D_E" 
    - "P6_10B - chr1 loss" -> "P6_10B"
    - "HyB_B - 2x" -> "HyB_B"
    """
    if pd.isna(cell_value) or cell_value == '':
        return None
    
    # Convert to string and trim
    str_val = str(cell_value).strip()
    
    # Handle special cases first
    if 'Col-0_' in str_val:
        match = re.search(r'Col-0_([A-Za-z0-9]+)', str_val)
        if match:
            return f"Col_0_{match.group(1)}"
    
    if 'Ler-D_' in str_val:
        match = re.search(r'Ler-D_([A-Za-z0-9]+)', str_val)
        if match:
            return f"Ler_D_{match.group(1)}"
    
    # Split by space or dash and find ID pattern
    parts = re.split(r'[\s-]+', str_val)
    
    for part in parts:
        # Look for patterns like P4_30A, HyB_A, etc.
        if re.match(r'^[A-Za-z0-9]+_[A-Za-z0-9]+$', part):
            return part
    
    return None


def load_order_file(order_path):
    """Load the order file and extract IDs from the first column."""
    try:
        # Try reading the Excel file
        df = pd.read_excel(order_path, header=None)
        order_ids = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        order_ids = [id for id in order_ids if id != '']
        
        print(f"✅ Loaded {len(order_ids)} IDs from order file")
        return order_ids
    
    except Exception as e:
        print(f"❌ Error reading order file: {e}")
        return None


def load_results_file(results_path):
    """Load the results file while preserving formatting."""
    try:
        # Load with pandas for data processing
        df = pd.read_excel(results_path, header=None)
        
        # Also load with openpyxl to preserve formatting
        wb = load_workbook(results_path)
        ws = wb.active
        
        print(f"✅ Loaded results file with {len(df)} rows and {len(df.columns)} columns")
        return df, wb, ws
    
    except Exception as e:
        print(f"❌ Error reading results file: {e}")
        return None, None, None


def create_id_mapping(df):
    """Create mapping from extracted IDs to row indices."""
    id_to_rows = {}
    extraction_log = []
    
    for idx, row in df.iterrows():
        # Extract ID from second column (index 1)
        extracted_id = extract_id(row.iloc[1] if len(row) > 1 else None)
        
        extraction_log.append({
            'row': idx + 1,
            'original': row.iloc[1] if len(row) > 1 else '',
            'extracted_id': extracted_id
        })
        
        if extracted_id:
            if extracted_id not in id_to_rows:
                id_to_rows[extracted_id] = []
            id_to_rows[extracted_id].append(idx)
    
    return id_to_rows, extraction_log


def reorder_dataframe(df, order_ids, id_to_rows):
    """Reorder the dataframe based on the order IDs."""
    reordered_indices = []
    used_indices = set()
    matching_log = []
    duplicate_row_indices = []  # Track which rows are duplicates
    
    # First pass: add rows in order
    for order_id in order_ids:
        if order_id in id_to_rows:
            row_indices = id_to_rows[order_id]
            for i, row_idx in enumerate(row_indices):
                reordered_indices.append(row_idx)
                used_indices.add(row_idx)
                
                # Mark duplicates (all matches except the first one)
                if i > 0:
                    duplicate_row_indices.append(len(reordered_indices) - 1)  # Position in reordered list
            
            matching_log.append({
                'order_id': order_id,
                'matches': len(row_indices),
                'found': True
            })
        else:
            matching_log.append({
                'order_id': order_id,
                'matches': 0,
                'found': False
            })
    
    # Second pass: add unmatched rows at the bottom
    unmatched_indices = []
    for idx in range(len(df)):
        if idx not in used_indices:
            reordered_indices.append(idx)
            unmatched_indices.append(idx)
    
    # Reorder the dataframe
    reordered_df = df.iloc[reordered_indices].reset_index(drop=True)
    
    return reordered_df, reordered_indices, matching_log, unmatched_indices, duplicate_row_indices


def save_reordered_file(reordered_df, original_wb, original_ws, reordered_indices, duplicate_row_indices, output_path):
    """Save the reordered file while preserving formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Alignment
        
        # Create a new workbook
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Merged_Analysis_Results"
        
        # Create yellow fill for duplicates
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Copy reordered data with full formatting preservation
        for new_row_idx, orig_row_idx in enumerate(reordered_indices):
            for col_idx in range(len(reordered_df.columns)):
                # Get original cell
                orig_cell = original_ws.cell(row=orig_row_idx + 1, column=col_idx + 1)
                
                # Create new cell
                new_cell = new_ws.cell(row=new_row_idx + 1, column=col_idx + 1)
                
                # Copy value
                new_cell.value = orig_cell.value
                
                # Copy formatting exactly like your Results Merger script
                if orig_cell.font:
                    new_cell.font = Font(
                        name=orig_cell.font.name,
                        size=orig_cell.font.size,
                        bold=orig_cell.font.bold,
                        italic=orig_cell.font.italic,
                        color=orig_cell.font.color
                    )
                
                # Apply yellow fill for duplicate rows, otherwise use original fill
                if new_row_idx in duplicate_row_indices:
                    new_cell.fill = yellow_fill
                elif orig_cell.fill:
                    new_cell.fill = PatternFill(
                        start_color=orig_cell.fill.start_color,
                        end_color=orig_cell.fill.end_color,
                        fill_type=orig_cell.fill.fill_type
                    )
                
                if orig_cell.border:
                    new_cell.border = Border(
                        left=orig_cell.border.left,
                        right=orig_cell.border.right,
                        top=orig_cell.border.top,
                        bottom=orig_cell.border.bottom
                    )
                
                if orig_cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=orig_cell.alignment.horizontal,
                        vertical=orig_cell.alignment.vertical,
                        wrap_text=orig_cell.alignment.wrap_text
                    )
        
        # Copy column widths and row heights if they exist
        if hasattr(original_ws, 'column_dimensions'):
            for col, dimension in original_ws.column_dimensions.items():
                new_ws.column_dimensions[col].width = dimension.width
                
        if hasattr(original_ws, 'row_dimensions'):
            for new_row_idx, orig_row_idx in enumerate(reordered_indices):
                orig_row_key = orig_row_idx + 1
                new_row_key = new_row_idx + 1
                if orig_row_key in original_ws.row_dimensions:
                    new_ws.row_dimensions[new_row_key].height = original_ws.row_dimensions[orig_row_key].height
        
        # Save the file
        new_wb.save(output_path)
        print(f"✅ Saved reordered file with preserved formatting to: {output_path}")
        print(f"💛 Highlighted {len(duplicate_row_indices)} duplicate rows in yellow")
        return True
        
    except Exception as e:
        print(f"⚠️  Could not preserve formatting: {e}")
        try:
            # Fallback: save with pandas
            reordered_df.to_excel(output_path, index=False, header=False)
            print(f"✅ Saved reordered file (basic format) to: {output_path}")
            return True
        except Exception as e2:
            print(f"❌ Error saving file: {e2}")
            return False


def print_summary(order_ids, matching_log, unmatched_indices, extraction_log):
    """Print a summary of the reordering process."""
    matched_count = sum(1 for log in matching_log if log['found'])
    multiple_matches = sum(1 for log in matching_log if log['matches'] > 1)
    total_matched_rows = sum(log['matches'] for log in matching_log)
    
    print("\n" + "="*60)
    print("📊 REORDERING SUMMARY")
    print("="*60)
    print(f"Order file IDs:           {len(order_ids)}")
    print(f"IDs found in results:     {matched_count}")
    print(f"IDs with multiple matches: {multiple_matches}")
    print(f"Total matched rows:       {total_matched_rows}")
    print(f"Unmatched rows:           {len(unmatched_indices)}")
    
    # Show some examples
    print(f"\n🔍 ID EXTRACTION EXAMPLES:")
    for log in extraction_log[:10]:
        if log['extracted_id']:
            print(f"  Row {log['row']:3d}: '{log['original'][:50]:<50}' → '{log['extracted_id']}'")
    
    # Show unmatched order IDs
    not_found = [log['order_id'] for log in matching_log if not log['found']]
    if not_found:
        print(f"\n❌ ORDER IDs NOT FOUND ({len(not_found)}):")
        for i, missing_id in enumerate(not_found[:10]):
            print(f"  {missing_id}")
        if len(not_found) > 10:
            print(f"  ... and {len(not_found) - 10} more")
    
    # Show multiple matches
    multiple = [log for log in matching_log if log['matches'] > 1]
    if multiple:
        print(f"\n🔄 IDs WITH MULTIPLE MATCHES ({len(multiple)}):")
        for log in multiple[:10]:
            print(f"  {log['order_id']}: {log['matches']} matches")
        if len(multiple) > 10:
            print(f"  ... and {len(multiple) - 10} more")


def main():
    parser = argparse.ArgumentParser(description='Reorder Excel file based on ID order')
    parser.add_argument('--order', '-o', 
                       help='Path to order Excel file',
                       default='/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Downloads/order.xlsx')
    parser.add_argument('--results', '-r',
                       help='Path to results Excel file', 
                       default='/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master\'s Thesis/Experiments/Aneuploidy Testing/Results/Merged_Analysis_Results.xlsx')
    parser.add_argument('--output', '-out',
                       help='Output file path',
                       default='/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master\'s Thesis/Experiments/Aneuploidy Testing/Results/Reordered_Merged_Analysis_Results.xlsx')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Show detailed processing information')
    
    args = parser.parse_args()
    
    print("🔄 Excel File Reordering Script")
    print("="*40)
    
    # Check if files exist
    order_path = Path(args.order)
    results_path = Path(args.results)
    
    if not order_path.exists():
        print(f"❌ Order file not found: {order_path}")
        print("   Please provide the correct path with --order")
        return 1
    
    if not results_path.exists():
        print(f"❌ Results file not found: {results_path}")
        print("   Please provide the correct path with --results")
        return 1
    
    print(f"📁 Order file:   {order_path}")
    print(f"📁 Results file: {results_path}")
    print(f"📁 Output file:  {args.output}")
    print()
    
    # Load files
    print("1️⃣ Loading order file...")
    order_ids = load_order_file(order_path)
    if order_ids is None:
        return 1
    
    print("2️⃣ Loading results file...")
    results_df, results_wb, results_ws = load_results_file(results_path)
    if results_df is None:
        return 1
    
    print("3️⃣ Extracting IDs from results file...")
    id_to_rows, extraction_log = create_id_mapping(results_df)
    print(f"   Found {len(id_to_rows)} unique IDs")
    
    print("4️⃣ Reordering data...")
    reordered_df, reordered_indices, matching_log, unmatched_indices, duplicate_row_indices = reorder_dataframe(
        results_df, order_ids, id_to_rows
    )
    
    print("5️⃣ Saving reordered file...")
    success = save_reordered_file(
        reordered_df, results_wb, results_ws, reordered_indices, duplicate_row_indices, args.output
    )
    
    if success:
        print_summary(order_ids, matching_log, unmatched_indices, extraction_log)
        print(f"\n🎉 Successfully reordered and saved to: {args.output}")
        return 0
    else:
        return 1


if __name__ == "__main__":
    sys.exit(main())
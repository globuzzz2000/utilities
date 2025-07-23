import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def merge_analysis_results():
    # Define the folder paths
    folders = [
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E7_B1_Initial Trial",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E8_B1_Batch1",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E8_B2_Batch2",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E8_B3_Batch2",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E8_B4_Batch2",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E9_B1_Redo_Control_WT",
        "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/JaW_E9_B2_LCO"
    ]
    
    # Create a new workbook for the merged results
    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Merged_Analysis_Results"
    
    header_written = False
    current_row = 1
    
    for folder in folders:
        file_path = os.path.join(folder, "Analysis_Results.xlsx")
        
        if not os.path.exists(file_path):
            print(f"Warning: File not found at {file_path}")
            continue
        
        try:
            # Load the workbook to preserve formatting
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Skip first two rows, start from row 3
            start_row = 3
            
            # If this is the first file, copy the header (row 3)
            if not header_written:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=start_row, column=col)
                    new_cell = merged_ws.cell(row=current_row, column=col)
                    
                    # Copy value
                    new_cell.value = cell.value
                    
                    # Copy formatting
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color
                        )
                    
                    if cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                            fill_type=cell.fill.fill_type
                        )
                    
                    if cell.border:
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom
                        )
                    
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text
                        )
                
                header_written = True
                current_row += 1
                start_row += 1  # Move to data rows
            
            # Copy data rows (starting from row 4 in source, or row after header in source)
            for row in range(start_row, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    new_cell = merged_ws.cell(row=current_row, column=col)
                    
                    # Copy value
                    new_cell.value = cell.value
                    
                    # Copy formatting
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color
                        )
                    
                    if cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                            fill_type=cell.fill.fill_type
                        )
                    
                    if cell.border:
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom
                        )
                    
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text
                        )
                
                current_row += 1
            
            print(f"Successfully processed: {folder}")
            
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            continue
    
    # Save the merged file
    output_path = "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Experiments/Aneuploidy Testing/Results/Merged_Analysis_Results.xlsx"
    merged_wb.save(output_path)
    print(f"Merged file saved as: {output_path}")

if __name__ == "__main__":
    merge_analysis_results()
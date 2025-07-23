#!/usr/bin/env python3
"""
Standalone script to assign chromosome and position columns to primer Excel files
based on exact matches in a FASTA file with BLAST fallback.

This script:
1. Reads an Excel file containing primer data
2. First tries exact string matching (including reverse complement) in FASTA
3. Falls back to BLAST for sequences not found by exact matching
4. Assigns chromosome and position based on the Forward Primer location
5. Updates the Gene column based on annotation file using the detected location
6. Outputs a new Excel file with updated position information and match quality indicators
7. Uses matching formatting from file_io.py for consistent output styling
"""

import pandas as pd
import os
import sys
import tempfile
import subprocess
import logging
import json
from pathlib import Path

# Optional import for Excel formatting (matching file_io.py)
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============= INPUT FILES - MODIFY THESE PATHS =============
EXCEL_FILE = "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Primer Design/Primers/Primers_Lyrata.xlsx"  # Path to Excel file
FASTA_FILE = "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Primer Design/Arenosa.fna"  # Path to FASTA file
ANNOTATION_FILE = "/Users/jakob/Library/Mobile Documents/com~apple~CloudDocs/Studium/Master/Master's Thesis/Primer Design/Arenosa.gff.gz"  # Path to GFF/GTF annotation file
USER_SETTINGS_PATH = "/Users/jakob/.ddPrimer/user_settings.json"  # Path to user settings JSON file
# ============================================================

# BLAST Configuration
BLAST_WORD_SIZE = 7
BLAST_EVALUE = 1000
BLAST_REWARD = 2
BLAST_PENALTY = -3
BLAST_GAPOPEN = 5
BLAST_GAPEXTEND = 2
BLAST_MAX_TARGET_SEQS = 10

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def get_blast_db_path():
    """
    Read the BLAST database path from the user settings JSON file.
    
    Returns:
        str: Path to BLAST database, or None if not found
    """
    try:
        if not os.path.exists(USER_SETTINGS_PATH):
            print(f"WARNING: User settings file not found: {USER_SETTINGS_PATH}")
            return None
            
        with open(USER_SETTINGS_PATH, 'r') as f:
            settings = json.load(f)
            
        blast_db_path = settings.get('blast_db_path')
        if not blast_db_path:
            print("WARNING: 'blast_db_path' not found in user settings file")
            return None
            
        print(f"Found BLAST database path in settings: {blast_db_path}")
        return blast_db_path
        
    except json.JSONDecodeError as e:
        print(f"WARNING: Invalid JSON in user settings file: {e}")
        return None
    except Exception as e:
        print(f"Error reading Excel with merged headers: {e}")
        print("Falling back to standard reading...")
        return

def read_excel_with_merged_headers(excel_path):
    """
    Read Excel file handling merged headers properly, especially for the Gene column.
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        DataFrame with properly named columns
    """
    try:
        if HAS_OPENPYXL:
            # First, read the raw data to understand the structure
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # Check for merged cells in the header area and get their values
            merged_ranges = worksheet.merged_cells.ranges
            header_row_1 = {}
            header_row_2 = {}
            
            # Read row 1 and row 2 headers
            for col_num in range(1, worksheet.max_column + 1):
                cell_1 = worksheet.cell(row=1, column=col_num)
                cell_2 = worksheet.cell(row=2, column=col_num)
                
                header_row_1[col_num] = cell_1.value
                header_row_2[col_num] = cell_2.value
            
            # Check for merged cells that span rows 1-2
            for merged_range in merged_ranges:
                if merged_range.min_row == 1 and merged_range.max_row == 2:
                    # This is a merged cell spanning both header rows
                    col_num = merged_range.min_col
                    merged_value = worksheet.cell(row=1, column=col_num).value
                    header_row_2[col_num] = merged_value  # Use the merged value as the column name
            
            workbook.close()
            
            # Now read with pandas using header=1, then fix column names
            df = pd.read_excel(excel_path, header=1)
            
            # Fix column names based on what we found
            new_columns = []
            for i, col in enumerate(df.columns, 1):
                if header_row_2.get(i):
                    new_columns.append(header_row_2[i])
                else:
                    new_columns.append(col)
            
            df.columns = new_columns
            return df
            
        else:
            # Fallback to regular pandas reading
            return pd.read_excel(excel_path, header=1)
            
    except Exception as e:
        print(f"Error reading Excel with merged headers: {e}")
        print("Falling back to standard reading...")
        try:
            # Try reading with header=1 first (skip the first row)
            return pd.read_excel(excel_path, header=1)
        except Exception as e2:
            print(f"Error reading with header=1: {e2}")
            print("Trying with header=0...")
            try:
                # Last resort: read with header=0 (use first row as headers)
                return pd.read_excel(excel_path, header=0)
            except Exception as e3:
                print(f"Error reading with header=0: {e3}")
                # Final fallback: read without headers and assign generic column names
                print("Reading without headers as final fallback...")
                df = pd.read_excel(excel_path, header=None)
                df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]
                return df

def process_excel_file(excel_path, fasta_path, annotation_path):
    """
    Process the Excel file and assign positions based on exact matches with BLAST fallback.
    Now includes proper column standardization and formatting matching file_io.py.
    Also updates Gene column based on annotation file.
    
    Args:
        excel_path: Path to the Excel file
        fasta_path: Path to the FASTA file
        annotation_path: Path to the annotation file
    """
    # Load the Excel file with proper merged header handling
    print(f"\nLoading Excel file: {excel_path}")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    df = read_excel_with_merged_headers(excel_path)
    print(f"Loaded Excel file with {len(df)} rows and {len(df.columns)} columns")
    
    # Display column names
    print("Available columns:")
    for i, col in enumerate(df.columns):
        print(f"  {i+1}. {col}")
    
    # Load the FASTA file
    genome_sequences = load_fasta(fasta_path)
    
    # Load the annotation file
    annotations = load_annotation_file(annotation_path)
    
    # Get BLAST database path (optional)
    blast_db_path = get_blast_db_path()
    if blast_db_path and os.path.exists(f"{blast_db_path}.nhr"):
        print(f"BLAST database available: {blast_db_path}")
        use_blast = True
    else:
        print("BLAST database not available - using exact matching only")
        use_blast = False
    
    # Find or create Chr and Location columns
    chr_col_idx = None
    loc_col_idx = None
    
    # Look for existing Chr and Location columns
    for i, col in enumerate(df.columns):
        if col == 'Chr':
            chr_col_idx = i
        elif col == 'Location':
            loc_col_idx = i
    
    # Add columns if they don't exist (at the end)
    if chr_col_idx is None:
        df['Chr'] = None
        chr_col_idx = len(df.columns) - 1
    if loc_col_idx is None:
        df['Location'] = None
        loc_col_idx = len(df.columns) - 1
    
    # Add match quality column if using BLAST
    if use_blast:
        df['Match_Quality'] = None
    
    # Process each row
    print(f"\nProcessing {len(df)} primer sets...")
    
    matches_found = 0
    no_matches = 0
    exact_matches = 0
    blast_matches = 0
    genes_updated = 0
    
    for idx, row in df.iterrows():
        # Get primer sequences - try multiple possible column names
        forward_primer = None
        reverse_primer = None
        probe = None
        
        # Try different column name variations
        for col in df.columns:
            col_lower = str(col).lower()
            if 'sequence' in col_lower and '(f)' in col_lower:
                forward_primer = row.get(col, '')
            elif 'sequence' in col_lower and '(r)' in col_lower:
                reverse_primer = row.get(col, '')
            elif 'sequence' in col_lower and '(p)' in col_lower:
                probe = row.get(col, '')
        
        # Fallback to exact column names
        if not forward_primer:
            forward_primer = row.get('Sequence (F)', '')
        if not reverse_primer:
            reverse_primer = row.get('Sequence (R)', '')
        if not probe:
            probe = row.get('Sequence (P)', '')
        
        # Get gene name - try multiple possible column names, including the properly read "Gene"
        gene_name = None
        original_gene_name = None
        for col in ['Gene', 'Unnamed: 0', 'Target', 'Name']:
            if col in df.columns:
                gene_value = row.get(col, f'Row_{idx+1}')
                if pd.notna(gene_value) and str(gene_value).strip():
                    original_gene_name = str(gene_value).strip()
                    gene_name = original_gene_name
                    break
        if not gene_name:
            gene_name = f'Row_{idx+1}'
            original_gene_name = gene_name
        
        print(f"\nProcessing {gene_name} (row {idx+1}):")
        
        # Find positions for all sequences with fallback
        forward_chr, forward_pos, forward_info, forward_match = find_sequence_position_with_fallback(
            forward_primer, genome_sequences, blast_db_path)
        reverse_chr, reverse_pos, reverse_info, reverse_match = find_sequence_position_with_fallback(
            reverse_primer, genome_sequences, blast_db_path)
        probe_chr, probe_pos, probe_info, probe_match = find_sequence_position_with_fallback(
            probe, genome_sequences, blast_db_path)
        
        # Check if ALL sequences are found
        all_found = True
        match_methods = []
        
        if forward_chr and forward_pos:
            print(f"  ✓ Found Forward Primer at {forward_chr}:{forward_pos} ({forward_match}, {forward_info})")
            match_methods.append(forward_match)
        else:
            print(f"  ✗ Forward Primer not found in genome")
            all_found = False
        
        if reverse_primer and str(reverse_primer).strip():  # Only check if reverse primer exists
            if reverse_chr and reverse_pos:
                print(f"  ✓ Found Reverse Primer at {reverse_chr}:{reverse_pos} ({reverse_match}, {reverse_info})")
                match_methods.append(reverse_match)
            else:
                print(f"  ✗ Reverse Primer not found in genome")
                all_found = False
        
        if probe and str(probe).strip():  # Only check if probe exists
            if probe_chr and probe_pos:
                print(f"  ✓ Found Probe at {probe_chr}:{probe_pos} ({probe_match}, {probe_info})")
                match_methods.append(probe_match)
            else:
                print(f"  ✗ Probe not found in genome")
                all_found = False
        
        # Only assign position if ALL sequences are found
        if all_found:
            df.at[idx, 'Chr'] = forward_chr
            df.at[idx, 'Location'] = forward_pos
            
            # Determine overall match quality
            if use_blast:
                if all('exact' in method for method in match_methods):
                    df.at[idx, 'Match_Quality'] = 'exact'
                    exact_matches += 1
                else:
                    # Use the worst match quality
                    worst_match = 'exact'
                    for method in match_methods:
                        if 'blast' in method:
                            worst_match = method
                            break
                    df.at[idx, 'Match_Quality'] = worst_match
                    blast_matches += 1
            
            # =================== NEW: ANNOTATION UPDATE LOGIC ===================
            # Update Gene column based on annotation
            if annotations:
                annotation_gene = find_gene_at_position(forward_chr, forward_pos, annotations)
                if annotation_gene:
                    # Ensure the 'Gene' column exists before trying to assign to it
                    if 'Gene' not in df.columns:
                        # If 'Gene' column doesn't exist, create it.
                        # This might happen if the input file has a different name for the gene column.
                        df['Gene'] = None 
                    
                    df.at[idx, 'Gene'] = annotation_gene
                    genes_updated += 1
                    
                    if annotation_gene != original_gene_name:
                        print(f"  ✓ Gene updated: '{original_gene_name}' → '{annotation_gene}'")
                    else:
                        print(f"  ✓ Gene annotation confirmed: '{annotation_gene}'")
                else:
                    print(f"  ! No gene annotation found at {forward_chr}:{forward_pos}")
            # ====================================================================

            matches_found += 1
            print(f"  ✓ ALL sequences found - Position assigned: {forward_chr}:{forward_pos}")
        else:
            no_matches += 1
            print(f"  ✗ NOT all sequences found - Marked as 'Not found'")
            df.at[idx, 'Chr'] = "Not found"
            # Leave Location as original value or empty, don't set to "Not found"
            if use_blast:
                df.at[idx, 'Match_Quality'] = "Not found"
    
    # Prepare the DataFrame for output with standardized column names
    df = prepare_output_dataframe(df)
    
    # Define final column order matching file_io.py standards
    columns = [
        "Gene",
        "Sequence (F)", "Tm (F)", "Penalty (F)", "dG (F)", "BLAST (F)",
        "Sequence (R)", "Tm (R)", "Penalty (R)", "dG (R)", "BLAST (R)",
        "Sequence (A)", "Length", "GC%", "dG (A)",
        "Chr", "Location"
    ]
    
    # Add probe columns if present
    if "Sequence (P)" in df.columns:
        probe_cols = ["Sequence (P)", "Tm (P)", "Penalty (P)", "dG (P)", "BLAST (P)"]
        # Insert probe columns after reverse primer columns
        try:
            idx_insert = columns.index("BLAST (R)") + 1
            for col in reversed(probe_cols):
                if col in df.columns:
                    columns.insert(idx_insert, col)
        except ValueError:
            for col in probe_cols:
                if col in df.columns:
                    columns.append(col)
    
    # Only select columns that actually exist in the DataFrame
    available_columns = [col for col in columns if col in df.columns]
    missing_columns = [col for col in columns if col not in df.columns]
    
    if missing_columns:
        print(f"Warning: Missing expected columns: {missing_columns}")
        print(f"Available columns: {df.columns.tolist()}")
    
    print(f"Final column selection: {available_columns}")
    df = df[available_columns]
    
    # Create output filename
    input_path = Path(excel_path)
    output_filename = f"{input_path.stem}_with_positions{input_path.suffix}"
    output_path = input_path.parent / output_filename
    
    # Save the updated Excel file with matching formatting
    print(f"\nSaving results to: {output_path}")
    try:
        format_excel_matching_file_io(df, str(output_path))
        print(f"Excel file saved with file_io.py formatting to: {output_path}")
    except Exception as e:
        print(f"Error applying file_io.py formatting: {str(e)}")
        print("Falling back to basic Excel export...")
        df.to_excel(output_path, index=False)
        print(f"Basic Excel file saved to: {output_path}")
    
    # Summary
    print(f"\n=== SUMMARY ===")
    print(f"Total primer sets processed: {len(df)}")
    print(f"Positions found: {matches_found}")
    if use_blast:
        print(f"  - Exact matches: {exact_matches}")
        print(f"  - BLAST matches: {blast_matches}")
    print(f"No matches: {no_matches}")
    # =================== NEW: ANNOTATION SUMMARY ===================
    if annotations:
        print(f"Genes updated from annotation: {genes_updated}")
    # ===============================================================
    print(f"Success rate: {matches_found/len(df)*100:.1f}%")
    print(f"\nOutput saved to: {output_path}")

def load_annotation_file(annotation_path):
    """
    Load annotation file (GFF/GTF format) and create a lookup structure.
    
    Args:
        annotation_path: Path to the annotation file
        
    Returns:
        Dictionary with chromosome as key and list of gene features as value
    """
    annotations = {}
    
    # =================== MODIFICATION: Handle None annotation_path ===================
    if not annotation_path or not os.path.exists(annotation_path):
        if annotation_path:  # Only print warning if a path was actually provided
            print(f"WARNING: Annotation file not found: {annotation_path}")
        return annotations
    # ===============================================================================
    
    print(f"Loading annotation file: {annotation_path}")
    
    # Determine if the file is gzipped
    is_gzipped = annotation_path.endswith('.gz')
    
    try:
        # Choose open method based on file type
        open_func = gzip.open if is_gzipped else open
        mode = 'rt' if is_gzipped else 'r'
        
        with open_func(annotation_path, mode) as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                
                # Skip comments and empty lines
                if not line or line.startswith('#'):
                    continue
                
                try:
                    # Parse GFF/GTF line
                    fields = line.split('\t')
                    if len(fields) < 9:
                        continue
                    
                    chromosome = fields[0]
                    feature_type = fields[2]
                    start = int(fields[3])
                    end = int(fields[4])
                    strand = fields[6]
                    attributes = fields[8]
                    
                    # Only process gene features
                    if feature_type.lower() not in ['gene']:
                        continue
                    
                    # Extract Name from attributes
                    gene_name = None
                    
                    # Parse attributes (handle both GFF and GTF formats)
                    # Priority: Name=, then gene_name=, then gene_id=
                    attrs_list = [a.strip() for a in attributes.split(';')]
                    
                    for attr in attrs_list:
                        if attr.startswith('Name='):
                            gene_name = attr.split('=', 1)[1]
                            break
                    
                    if not gene_name:
                        for attr in attrs_list:
                            if attr.startswith('gene_name'):
                                gene_name = attr.split(maxsplit=1)[1]
                                break
                    
                    if not gene_name:
                        for attr in attrs_list:
                            if attr.startswith('gene_id'):
                                gene_name = attr.split(maxsplit=1)[1]
                                break

                    # If no Name found, skip this entry
                    if not gene_name:
                        continue
                    
                    # Remove quotes if present
                    gene_name = gene_name.strip('"\'')
                    
                    # Add to annotations
                    if chromosome not in annotations:
                        annotations[chromosome] = []
                    
                    annotations[chromosome].append({
                        'start': start,
                        'end': end,
                        'name': gene_name,
                        'strand': strand,
                        'type': feature_type
                    })
                    
                except (ValueError, IndexError) as e:
                    print(f"Warning: Could not parse line {line_num} in annotation file: {e}")
                    continue
        
        # Sort annotations by start position for efficient searching
        for chromosome in annotations:
            annotations[chromosome].sort(key=lambda x: x['start'])
        
        total_genes = sum(len(genes) for genes in annotations.values())
        print(f"Loaded {total_genes} gene annotations from {len(annotations)} chromosomes")
        
        # Show some statistics
        # Use sorted keys for consistent output
        sorted_chroms = sorted(annotations.keys())
        for chrom in sorted_chroms[:5]:
            print(f"  {chrom}: {len(annotations[chrom])} genes")
        if len(annotations) > 5:
            print(f"  ... and {len(annotations) - 5} more chromosomes")
        
    except Exception as e:
        print(f"ERROR: Failed to load annotation file: {e}")
        return {}
    
    return annotations

def find_gene_at_position(chromosome, position, annotations):
    """
    Find the gene that contains the given position.
    
    Args:
        chromosome: Chromosome name
        position: Position on chromosome (1-based)
        annotations: Annotation dictionary from load_annotation_file
        
    Returns:
        Gene name if found, None otherwise
    """
    if not annotations or chromosome not in annotations:
        return None
    
    genes = annotations[chromosome]
    
    # Binary search would be more efficient for very large lists,
    # but a linear scan is simple and effective for moderately sized lists.
    # Since genes are sorted, we can also break early.
    for gene in genes:
        if gene['start'] <= position <= gene['end']:
            return gene['name']
        elif gene['start'] > position:
            # Since genes are sorted by start position, we can stop searching
            break
    
    return None

def reverse_complement(sequence):
    """
    Return the reverse complement of a DNA sequence.
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        Reverse complement sequence
    """
    if not sequence or pd.isna(sequence):
        return None
    
    complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G',
                  'a': 't', 't': 'a', 'g': 'c', 'c': 'g',
                  'N': 'N', 'n': 'n'} # Handle ambiguous bases
    
    sequence = str(sequence).strip()
    try:
        rev_comp = ''.join(complement.get(base, 'N') for base in reversed(sequence))
        return rev_comp.upper()
    except KeyError as e:
        print(f"Warning: Invalid DNA base '{e.args[0]}' in sequence: {sequence}")
        return None

def load_fasta(filepath):
    """
    Load sequences from a FASTA file into a dictionary.
    
    Args:
        filepath: Path to the FASTA file
        
    Returns:
        Dictionary mapping sequence headers to sequences
    """
    sequences = {}
    name = None
    seq_chunks = []
    
    print(f"Loading FASTA file: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"FASTA file not found: {filepath}")
        
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name:
                    sequences[name] = "".join(seq_chunks).upper()
                name = line[1:].split()[0]  # Take only the first part of header
                seq_chunks = []
            else:
                seq_chunks.append(line)
                
        # Handle the last sequence
        if name:
            sequences[name] = "".join(seq_chunks).upper()
            
    print(f"Loaded {len(sequences)} sequences from FASTA file")
    # Sort by sequence name for consistent output
    for seq_name in sorted(sequences.keys()):
        seq = sequences[seq_name]
        print(f"  {seq_name}: {len(seq):,} bp")
    
    return sequences

def find_exact_position(sequence, genome_sequences):
    """
    Find the exact position of a sequence in the genome.
    Searches for both forward and reverse complement sequences.
    
    Args:
        sequence: DNA sequence to search for
        genome_sequences: Dictionary of chromosome sequences
        
    Returns:
        Tuple of (chromosome, position, strand, match_type) or (None, None, None, None) if not found
        strand is '+' for forward match, '-' for reverse complement match
        match_type is 'exact' for perfect matches
    """
    if not sequence or pd.isna(sequence):
        return None, None, None, None
    
    sequence = str(sequence).upper().strip()
    rev_comp_seq = reverse_complement(sequence)
    
    for chromosome, genome_seq in genome_sequences.items():
        # Search for forward sequence
        pos = genome_seq.find(sequence)
        if pos != -1:
            return chromosome, pos + 1, '+', 'exact'  # Convert to 1-based coordinates
        
        # Search for reverse complement
        if rev_comp_seq:
            pos = genome_seq.find(rev_comp_seq)
            if pos != -1:
                return chromosome, pos + 1, '-', 'exact'  # Convert to 1-based coordinates
    
    return None, None, None, None

def blast_sequence_for_position(seq, db_path):
    """
    Run BLASTn for a sequence and return the best match position.
    
    Args:
        seq: DNA sequence to BLAST against database
        db_path: BLAST database path
        
    Returns:
        Tuple of (chromosome, position, evalue, match_type) or (None, None, None, None) if not found
    """
    if not seq or not isinstance(seq, str) or not seq.strip():
        return None, None, None, None

    tmp_filename = None
    try:
        # Create temporary file for query sequence
        temp_dir = tempfile.gettempdir()
        with tempfile.NamedTemporaryFile(mode="w+", delete=False, dir=temp_dir, suffix=".fasta") as tmp_query:
            tmp_query.write(f">seq\n{seq}\n")
            tmp_query.flush()
            tmp_filename = tmp_query.name

        # Execute BLASTn command with detailed output format
        result = subprocess.run(
            [
                "blastn",
                "-task", "blastn-short",
                "-db", db_path,
                "-query", tmp_filename,
                "-word_size", str(BLAST_WORD_SIZE),
                "-evalue", str(BLAST_EVALUE),
                "-reward", str(BLAST_REWARD),
                "-penalty", str(BLAST_PENALTY),
                "-gapopen", str(BLAST_GAPOPEN),
                "-gapextend", str(BLAST_GAPEXTEND),
                "-max_target_seqs", str(BLAST_MAX_TARGET_SEQS),
                "-outfmt", "6 sseqid sstart send evalue bitscore length pident"
            ],
            text=True,
            capture_output=True
        )
        
        if result.returncode != 0:
            logger.error(f"BLAST execution failed for sequence {seq[:20]}...")
            logger.error(f"BLAST stderr: {result.stderr}")
            return None, None, None, None

        # Parse BLAST output
        lines = result.stdout.strip().split("\n")
        if not lines or not lines[0].strip():
            return None, None, None, None

        # Get the best hit (first line, already sorted by e-value)
        best_hit = lines[0].strip().split("\t")
        if len(best_hit) < 7:
            return None, None, None, None

        chromosome = best_hit[0]
        start_pos = int(best_hit[1])
        end_pos = int(best_hit[2])
        evalue = float(best_hit[3])
        percent_identity = float(best_hit[6])
        
        # Use the start position (5' end of the match)
        position = min(start_pos, end_pos)  # In case of reverse complement matches
        
        # Determine match type based on percent identity
        if percent_identity == 100.0:
            match_type = 'blast_perfect'
        else:
            match_type = f'blast_{percent_identity:.1f}%'
        
        return chromosome, position, evalue, match_type
        
    except FileNotFoundError:
        logger.error("`blastn` command not found. Please ensure BLAST+ is installed and in your system's PATH.")
        # Exit or handle gracefully
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error in BLAST for sequence {seq[:20]}...: {str(e)}")
        return None, None, None, None
        
    finally:
        # Clean up temporary file
        if tmp_filename and os.path.exists(tmp_filename):
            try:
                os.remove(tmp_filename)
            except OSError as e:
                logger.debug(f"Failed to remove temp file {tmp_filename}: {e}")

def find_sequence_position_with_fallback(sequence, genome_sequences, blast_db_path):
    """
    Find sequence position using exact matching first, then BLAST fallback.
    
    Args:
        sequence: DNA sequence to search for
        genome_sequences: Dictionary of chromosome sequences
        blast_db_path: Path to BLAST database (can be None)
        
    Returns:
        Tuple of (chromosome, position, additional_info, match_type)
    """
    # First try exact matching
    chr_exact, pos_exact, strand, match_type = find_exact_position(sequence, genome_sequences)
    
    if chr_exact and pos_exact:
        return chr_exact, pos_exact, f"strand:{strand}", match_type
    
    # If exact matching failed and BLAST is available, try BLAST
    if blast_db_path and os.path.exists(f"{blast_db_path}.nhr"):
        chr_blast, pos_blast, evalue, match_type = blast_sequence_for_position(sequence, blast_db_path)
        
        if chr_blast and pos_blast:
            return chr_blast, pos_blast, f"E-val:{evalue:.2e}", match_type
    
    return None, None, None, None

def calculate_gc(sequence):
    """
    Calculate GC% for a DNA sequence (matching file_io.py FilterProcessor method).
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        GC percentage as float, or None if sequence is invalid
    """
    if not sequence or pd.isna(sequence):
        return None
    
    sequence = str(sequence).upper().strip()
    if not sequence:
        return None
    
    gc_count = sequence.count('G') + sequence.count('C')
    total_count = len(sequence)
    
    if total_count == 0:
        return None
    
    return round((gc_count / total_count) * 100, 1)

def format_excel_matching_file_io(df, output_file):
    """
    Save DataFrame to Excel with comprehensive formatting matching file_io.py exactly.
    
    Args:
        df: DataFrame with primer results
        output_file: Path to save the formatted Excel file
        
    Returns:
        Path to the saved Excel file
    """
    try:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is not available")
            
        # First, save with pandas to get a basic Excel file
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Now open the file for formatting
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Create a new row for our custom headers
        worksheet.insert_rows(1)
        max_row += 1
        
        # Create styles
        header_font = Font(bold=True)
        sequence_fill = PatternFill(
            start_color='D9D9D9',
            end_color='D9D9D9',
            fill_type='solid'
        )
        centered_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Create border styles for group separators
        thin_border_left = Border(left=Side(style='thin'))
        thin_border_right = Border(right=Side(style='thin'))
        thin_border_both = Border(left=Side(style='thin'), right=Side(style='thin'))
        
        # Clear all default borders first
        no_border = Border()
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = no_border
        
        column_map = {}
        header_texts = []
        
        # Apply basic formatting
        for col_num in range(1, max_col + 1):
            col_letter = get_column_letter(col_num)
            
            # Set header row formatting
            cell2 = worksheet.cell(row=2, column=col_num)
            cell2.font = header_font
            cell2.alignment = centered_alignment
            
            # Handle Gene column
            if cell2.value == "Gene":
                cell1 = worksheet.cell(row=1, column=col_num)
                cell1.value = "Gene"
                cell1.font = header_font
                cell1.alignment = centered_alignment
                worksheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)
                # Set wider width for Gene column
                worksheet.column_dimensions[col_letter].width = 20
            else:
                # Set narrower width for non-Gene, non-sequence columns
                header_text = cell2.value
                if header_text and ("Sequence" in str(header_text)):
                    worksheet.column_dimensions[col_letter].width = 25
                else:
                    worksheet.column_dimensions[col_letter].width = 10

            header_text = cell2.value
            header_texts.append(header_text)
            column_map[header_text] = col_num
            
            # Apply number formatting based on column type
            for row_num in range(3, max_row + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Apply sequence fill to sequence columns
                if header_text and ("Sequence" in str(header_text)):
                    cell.fill = sequence_fill
                
                # Apply number formatting
                if header_text and "BLAST" in str(header_text):
                    # Scientific notation for BLAST columns (e.g., 1.23E-05)
                    cell.number_format = '0.00E+00'
                    cell.alignment = centered_alignment
                elif header_text and any(col_type in str(header_text) for col_type in ["Tm (", "Penalty (", "dG ("]):
                    # One decimal place for Tm, Penalty, and dG columns
                    cell.number_format = '0.0'
                    cell.alignment = centered_alignment
                elif header_text == "Length":
                    # Integer format for Length
                    cell.number_format = '0'
                    cell.alignment = centered_alignment
                elif header_text == "GC%":
                    # One decimal place for GC%
                    cell.number_format = '0.0'
                    cell.alignment = centered_alignment
                else:
                    # Handle "No suitable primers found" cells and other text
                    if cell.value == "No suitable primers found":
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = centered_alignment
        
        # Freeze panes
        worksheet.freeze_panes = 'B3'
        
        # Group columns with updated header names using position indicators
        header_groups = {
            "Gene": ["Gene"],
            "Forward Primer": ["Sequence (F)", "Tm (F)", "Penalty (F)", "dG (F)", "BLAST (F)"],
            "Reverse Primer": ["Sequence (R)", "Tm (R)", "Penalty (R)", "dG (R)", "BLAST (R)"],
            "Probe": ["Sequence (P)", "Tm (P)", "Penalty (P)", "dG (P)", "BLAST (P)"],
            "Amplicon": ["Sequence (A)", "Length", "GC%", "dG (A)"],
            "Location": ["Chr", "Location", "Match_Quality"]
        }
        
        # Track group boundaries for border application
        group_boundaries = []
        
        for group_name, headers in header_groups.items():
            existing_headers = [h for h in headers if h in header_texts]
            
            if not existing_headers:
                continue
                
            col_indices = [column_map[h] for h in existing_headers]
            
            if col_indices:
                start_col = min(col_indices)
                end_col = max(col_indices)
                
                # Store group boundaries
                group_boundaries.append((start_col, end_col))
                
                group_cell = worksheet.cell(row=1, column=start_col)
                group_cell.value = group_name
                group_cell.font = header_font
                group_cell.alignment = centered_alignment
                
                if start_col != end_col:
                    merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
                    try:
                        worksheet.merge_cells(merge_range)
                    except Exception as e:
                        logger.warning(f"Could not merge range {merge_range}: {str(e)}")
        
        # Apply borders to group boundaries
        for start_col, end_col in group_boundaries:
            # Apply left border to start of group
            for row_num in range(1, max_row + 1):
                cell = worksheet.cell(row=row_num, column=start_col)
                cell.border = Border(
                    left=Side(style='thin'),
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    right=cell.border.right
                )
            
            # Apply right border to end of group
            for row_num in range(1, max_row + 1):
                cell = worksheet.cell(row=row_num, column=end_col)
                cell.border = Border(
                    left=cell.border.left,
                    right=Side(style='thin'),
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
        
        # Add medium border around the entire populated table
        medium_border_side = Side(style='medium')
        
        # Top border
        for col_num in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.border = Border(
                top=medium_border_side,
                left=cell.border.left,
                right=cell.border.right,
                bottom=cell.border.bottom
            )
        
        # Bottom border
        for col_num in range(1, max_col + 1):
            cell = worksheet.cell(row=max_row, column=col_num)
            cell.border = Border(
                bottom=medium_border_side,
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top
            )
        
        # Left border
        for row_num in range(1, max_row + 1):
            cell = worksheet.cell(row=row_num, column=1)
            cell.border = Border(
                left=medium_border_side,
                top=cell.border.top,
                right=cell.border.right,
                bottom=cell.border.bottom
            )
        
        # Right border
        for row_num in range(1, max_row + 1):
            cell = worksheet.cell(row=row_num, column=max_col)
            cell.border = Border(
                right=medium_border_side,
                left=cell.border.left,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
        
        workbook.save(output_file)
        return output_file
            
    except ImportError:
        logger.warning("openpyxl not available, falling back to standard Excel export")
    except Exception as e:
        error_msg = f"Error applying Excel formatting: {str(e)}"
        logger.error(error_msg)
        logger.warning("Falling back to standard Excel export")
        logger.debug(f"Error details: {str(e)}", exc_info=True)
        
    # Basic fallback save
    try:
        df.to_excel(output_file, index=False)
        logger.info(f"Excel file saved to: {output_file} (without formatting)")
        return output_file
    except Exception as e:
        error_msg = f"Failed to save Excel file {output_file}: {str(e)}"
        logger.error(error_msg)
        logger.debug(f"Error details: {str(e)}", exc_info=True)
        raise Exception(error_msg) from e

def prepare_output_dataframe(df):
    """
    Prepare DataFrame for output by standardizing column names and adding derived fields.
    This matches the _prepare_output_dataframe function from file_io.py.
    
    Args:
        df: Input DataFrame with primer records
        
    Returns:
        DataFrame with properly formatted columns for output
    """
    output_df = df.copy()
    
    # 1. AMPLICON LENGTH: Create 'Length' from 'Amplicon Length' if it doesn't exist
    if 'Length' not in output_df.columns:
        if 'Amplicon Length' in output_df.columns:
            output_df['Length'] = output_df['Amplicon Length']
        elif 'Sequence (A)' in output_df.columns:
            output_df['Length'] = output_df['Sequence (A)'].apply(lambda x: len(str(x)) if pd.notna(x) else 0)

    # 2. Calculate GC% for amplicons if not present
    if 'GC%' not in output_df.columns and 'Sequence (A)' in output_df.columns:
        logger.debug("Calculating GC% for amplicon sequences")
        output_df['GC%'] = output_df['Sequence (A)'].apply(calculate_gc)

    # 3. LOCATION: Use only 'Start' coordinate if available
    if 'Start' in output_df.columns:
        output_df['Location'] = output_df['Start'].apply(
            lambda start: str(int(start)) if pd.notna(start) else ""
        )
    
    return output_df


def main():
    """Main function to run the position mapping script."""
    print("="*60)
    print("Excel Primer Position Mapper")
    print("="*60)
    
    try:
        # Check if input files exist
        if not os.path.exists(EXCEL_FILE):
            print(f"ERROR: Excel file not found: {EXCEL_FILE}")
            print("Please update the EXCEL_FILE variable at the top of this script.")
            sys.exit(1)
        
        if not os.path.exists(FASTA_FILE):
            print(f"ERROR: FASTA file not found: {FASTA_FILE}")
            print("Please update the FASTA_FILE variable at the top of this script.")
            sys.exit(1)
        
        # =================== MODIFICATION: Handle missing annotation file gracefully ===================
        annotation_path = None # Default to None
        if os.path.exists(ANNOTATION_FILE):
            annotation_path = ANNOTATION_FILE
        else:
            print(f"WARNING: Annotation file not found: {ANNOTATION_FILE}")
            print("Gene names will not be updated based on position.")
            print("To enable this feature, update the ANNOTATION_FILE variable.")
        # ==============================================================================================

        # Process the files
        process_excel_file(EXCEL_FILE, FASTA_FILE, annotation_path)
        
    except Exception as e:
        print(f"\nAn unexpected error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    # Add gzip import for gzipped annotation files
    import gzip 
    main()